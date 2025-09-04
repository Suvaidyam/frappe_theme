import frappe
from openpyxl import load_workbook, Workbook
import frappe
import io
import os
import json
import uuid 
import datetime
from openpyxl.utils.cell import get_column_letter,coordinate_from_string, column_index_from_string


def load_excel_from_private(file_url):
    file_path = os.path.join(frappe.get_site_path(), file_url.lstrip("/"))
    if not os.path.exists(file_path):
        frappe.throw(f"File {file_url} not found in private/files")
    wb = load_workbook(file_path, data_only=False)
    return wb


def rgb_to_hex(rgb_value):
    """Convert RGB value to hex format"""
    if not rgb_value:
        return None
    rgb_str = str(rgb_value)
    if len(rgb_str) == 8:  # ARGB format
        return f"#{rgb_str[2:].upper()}"
    elif len(rgb_str) == 6:  # RGB format
        return f"#{rgb_str.upper()}"
    return None


def extract_complete_cell_style(cell):
    """Extract ALL possible cell styling properties"""
    style = {}

    # ====== FILL/BACKGROUND ======
    if cell.fill:
        bg_info = {}
        if hasattr(cell.fill, 'start_color') and cell.fill.start_color:
            if hasattr(cell.fill.start_color, 'rgb') and cell.fill.start_color.rgb:
                bg_hex = rgb_to_hex(cell.fill.start_color.rgb)
                if bg_hex and bg_hex != "#000000":
                    bg_info["rgb"] = bg_hex
        
        # Pattern type
        if hasattr(cell.fill, 'fill_type') and cell.fill.fill_type:
            bg_info["pattern"] = cell.fill.fill_type
            
        # Gradient fills
        if hasattr(cell.fill, 'end_color') and cell.fill.end_color:
            if hasattr(cell.fill.end_color, 'rgb') and cell.fill.end_color.rgb:
                end_hex = rgb_to_hex(cell.fill.end_color.rgb)
                if end_hex:
                    bg_info["endColor"] = end_hex
        
        if bg_info:
            style["bg"] = bg_info

    # ====== FONT PROPERTIES ======
    if cell.font:
        font_info = {}
        
        # Font color
        if cell.font.color and hasattr(cell.font.color, 'rgb') and cell.font.color.rgb:
            font_hex = rgb_to_hex(cell.font.color.rgb)
            if font_hex:
                font_info["rgb"] = font_hex
        
        # Font formatting
        if cell.font.bold is True:
            style["bl"] = 1
        if cell.font.italic is True:
            style["it"] = 1
        if cell.font.underline and cell.font.underline != 'none':
            style["ul"] = 1
        if cell.font.strike is True:
            style["st"] = 1
            
        # Font size and family
        if cell.font.size:
            style["fs"] = int(cell.font.size)
        if cell.font.name:
            style["ff"] = cell.font.name
            
        # Font scheme
        if hasattr(cell.font, 'scheme') and cell.font.scheme:
            font_info["scheme"] = cell.font.scheme
            
        if font_info:
            style["cl"] = font_info

    # ====== ALIGNMENT ======
    if cell.alignment:
        # Horizontal alignment
        h_align = cell.alignment.horizontal
        if h_align == "center":
            style["ht"] = 2
        elif h_align == "right":
            style["ht"] = 3
        elif h_align == "left":
            style["ht"] = 1
        elif h_align == "justify":
            style["ht"] = 4
        
        # Vertical alignment
        v_align = cell.alignment.vertical
        if v_align == "center":
            style["vt"] = 2
        elif v_align == "top":
            style["vt"] = 1
        elif v_align == "bottom":
            style["vt"] = 3
        elif v_align == "justify":
            style["vt"] = 4
        
        # Text wrapping
        if cell.alignment.wrap_text is True:
            style["tb"] = 2  # wrap text
            
        # Text rotation
        if cell.alignment.text_rotation:
            style["tr"] = {"a": cell.alignment.text_rotation, "v": 0}
            
        # Indent
        if cell.alignment.indent:
            style["in"] = cell.alignment.indent

    # ====== BORDERS ======
    if cell.border:
        border = {}
        
        # Top border
        if cell.border.top and cell.border.top.style:
            top_border = {"s": 1}
            if cell.border.top.color and hasattr(cell.border.top.color, 'rgb') and cell.border.top.color.rgb:
                border_hex = rgb_to_hex(cell.border.top.color.rgb)
                if border_hex:
                    top_border["cl"] = {"rgb": border_hex}
            # Border style mapping
            if cell.border.top.style == "thin":
                top_border["s"] = 1
            elif cell.border.top.style == "medium":
                top_border["s"] = 2
            elif cell.border.top.style == "thick":
                top_border["s"] = 3
            elif cell.border.top.style == "double":
                top_border["s"] = 4
            border["t"] = top_border

        # Bottom border
        if cell.border.bottom and cell.border.bottom.style:
            bottom_border = {"s": 1}
            if cell.border.bottom.color and hasattr(cell.border.bottom.color, 'rgb') and cell.border.bottom.color.rgb:
                border_hex = rgb_to_hex(cell.border.bottom.color.rgb)
                if border_hex:
                    bottom_border["cl"] = {"rgb": border_hex}
            if cell.border.bottom.style == "thin":
                bottom_border["s"] = 1
            elif cell.border.bottom.style == "medium":
                bottom_border["s"] = 2
            elif cell.border.bottom.style == "thick":
                bottom_border["s"] = 3
            elif cell.border.bottom.style == "double":
                bottom_border["s"] = 4
            border["b"] = bottom_border

        # Left border
        if cell.border.left and cell.border.left.style:
            left_border = {"s": 1}
            if cell.border.left.color and hasattr(cell.border.left.color, 'rgb') and cell.border.left.color.rgb:
                border_hex = rgb_to_hex(cell.border.left.color.rgb)
                if border_hex:
                    left_border["cl"] = {"rgb": border_hex}
            if cell.border.left.style == "thin":
                left_border["s"] = 1
            elif cell.border.left.style == "medium":
                left_border["s"] = 2
            elif cell.border.left.style == "thick":
                left_border["s"] = 3
            elif cell.border.left.style == "double":
                left_border["s"] = 4
            border["l"] = left_border

        # Right border
        if cell.border.right and cell.border.right.style:
            right_border = {"s": 1}
            if cell.border.right.color and hasattr(cell.border.right.color, 'rgb') and cell.border.right.color.rgb:
                border_hex = rgb_to_hex(cell.border.right.color.rgb)
                if border_hex:
                    right_border["cl"] = {"rgb": border_hex}
            if cell.border.right.style == "thin":
                right_border["s"] = 1
            elif cell.border.right.style == "medium":
                right_border["s"] = 2
            elif cell.border.right.style == "thick":
                right_border["s"] = 3
            elif cell.border.right.style == "double":
                right_border["s"] = 4
            border["r"] = right_border

        if border:
            style["bd"] = border

    # ====== NUMBER FORMAT ======
    if cell.number_format and cell.number_format != "General":
        style["nf"] = {"t": "n", "v": cell.number_format}

    # ====== PROTECTION ======
    if cell.protection:
        if cell.protection.locked is False:
            style["lo"] = 0
        if cell.protection.hidden is True:
            style["hi"] = 1

    return style if style else None


def extract_complete_cell_value(cell):
    """Extract ALL possible cell value properties"""
    value_obj = {}

    # ====== FORMULA ======
    if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
        value_obj["f"] = cell.value
        # Try to get calculated value
        try:
            if hasattr(cell, 'calculated_value') and cell.calculated_value is not None:
                value_obj["v"] = cell.calculated_value
            else:
                value_obj["v"] = 0
        except:
            value_obj["v"] = 0
    else:
        # ====== VALUE TYPES ======
        if isinstance(cell.value, datetime.datetime):
            value_obj["v"] = cell.value.strftime("%Y-%m-%d %H:%M:%S")
            value_obj["t"] = "dt"
        elif isinstance(cell.value, datetime.date):
            value_obj["v"] = cell.value.strftime("%Y-%m-%d")
            value_obj["t"] = "d"
        elif isinstance(cell.value, datetime.time):
            value_obj["v"] = cell.value.strftime("%H:%M:%S")
            value_obj["t"] = "t"
        elif isinstance(cell.value, bool):
            value_obj["v"] = 1 if cell.value else 0
            value_obj["t"] = "b"
        elif isinstance(cell.value, (int, float)):
            value_obj["v"] = cell.value
            value_obj["t"] = "n"
        elif isinstance(cell.value, str):
            value_obj["v"] = cell.value
            value_obj["t"] = "s"
        elif cell.value is None:
            value_obj["v"] = ""
        else:
            value_obj["v"] = str(cell.value) if cell.value is not None else ""

    # ====== HYPERLINKS ======
    if cell.hyperlink:
        if hasattr(cell.hyperlink, 'target') and cell.hyperlink.target:
            value_obj["l"] = {
                "Target": cell.hyperlink.target,
                "Display": cell.hyperlink.display or cell.value or cell.hyperlink.target
            }
        elif hasattr(cell.hyperlink, 'location') and cell.hyperlink.location:
            value_obj["l"] = {
                "Target": f"#{cell.hyperlink.location}",
                "Display": cell.hyperlink.display or cell.value or cell.hyperlink.location
            }

    # ====== COMMENTS/NOTES ======
    if cell.comment:
        value_obj["ct"] = {
            "t": cell.comment.text,
            "a": cell.comment.author if hasattr(cell.comment, 'author') else ""
        }

    # ====== DATA VALIDATION ======
    if hasattr(cell, 'data_validation') and cell.data_validation:
        dv = cell.data_validation
        validation = {}
        if hasattr(dv, 'type') and dv.type:
            validation["type"] = dv.type
        if hasattr(dv, 'formula1') and dv.formula1:
            validation["formula1"] = str(dv.formula1)
        if hasattr(dv, 'formula2') and dv.formula2:
            validation["formula2"] = str(dv.formula2)
        if validation:
            value_obj["dv"] = validation

    return value_obj if value_obj else None


def is_row_hidden(sheet, row_idx):
    """Enhanced row hidden detection with detailed logging"""
    row_dimension = sheet.row_dimensions.get(row_idx)
    
    if row_dimension:
        # Check all possible hidden indicators
        hidden_property = getattr(row_dimension, 'hidden', False)
        height = getattr(row_dimension, 'height', None)
        outline_level = getattr(row_dimension, 'outline_level', 0)
        collapsed = getattr(row_dimension, 'collapsed', False)
        
        # print(f"Row {row_idx}: hidden={hidden_property}, height={height}, outline_level={outline_level}, collapsed={collapsed}")
        
        # Method 1: Direct hidden property
        if hidden_property is True:
            return True
        
        # Method 2: Height-based hiding
        if height is not None and height <= 0.1:
            return True
        
        # Method 3: Grouped/Outlined rows that are collapsed
        if outline_level > 0 and collapsed:
            return True
    
    return False
 
def is_column_hidden(sheet, col_idx):
    """Enhanced column hidden detection with detailed logging"""
    col_letter = get_column_letter(col_idx)
    col_dimension = sheet.column_dimensions.get(col_letter)
    
    if col_dimension:
        # Check all possible hidden indicators
        hidden_property = getattr(col_dimension, 'hidden', False)
        width = getattr(col_dimension, 'width', None)
        outline_level = getattr(col_dimension, 'outline_level', 0)
        collapsed = getattr(col_dimension, 'collapsed', False)
        
        # print(f"Col {col_letter}: hidden={hidden_property}, width={width}, outline_level={outline_level}, collapsed={collapsed}")
        
        # Method 1: Direct hidden property
        if hidden_property is True:
            return True
        
        # Method 2: Width-based hiding
        if width is not None and width <= 0.01:
            return True
        
        # Method 3: Grouped/Outlined columns that are collapsed
        if outline_level > 0 and collapsed:
            return True
    
    return False

def get_freeze_info(sheet):
    if sheet.freeze_panes:
        fp = sheet.freeze_panes
        if isinstance(fp, str):  
            # e.g. "C5"
            col, row = coordinate_from_string(fp)
            row_idx = row
            col_idx = column_index_from_string(col)
        else:  
            # OpenPyXL Cell object
            row_idx = fp.row
            col_idx = fp.col_idx

        return {
            "startRow": row_idx - 1 if row_idx else -1,
            "startColumn": col_idx - 1 if col_idx else -1,
            "ySplit": row_idx - 1 if row_idx else 0,
            "xSplit": col_idx - 1 if col_idx else 0,
        }
    else:
        return {"startRow": -1, "startColumn": -1, "ySplit": 0, "xSplit": 0}
    



#  ========= Main API Function ========
@frappe.whitelist(allow_guest=True)
def excel():
    """Version optimized for UniverJS with proper column hiding/collapsing"""
    wb = load_excel_from_private("/private/files/data.xlsx")
    
    sheets = {}
    sheet_order = []
    
    for sheet in wb.worksheets:
        sheet_id = str(uuid.uuid4())
        sheet_order.append(sheet_id)
        
        print(f"\n=== Processing Sheet: {sheet.title} ===")
        
        # ====== DETECT HIDDEN ELEMENTS ======
        hidden_rows_set = set()
        hidden_cols_set = set()
        
        for row_idx in range(1, sheet.max_row + 1):
            if is_row_hidden(sheet, row_idx):
                hidden_rows_set.add(row_idx - 1) 
                
        for col_idx in range(1, sheet.max_column + 1):
            if is_column_hidden(sheet, col_idx):
                hidden_cols_set.add(col_idx - 1) 
                col_letter = get_column_letter(col_idx)
        
        # print(f"Hidden rows (0-based): {sorted(hidden_rows_set)}")
        # print(f"Hidden columns (0-based): {sorted(hidden_cols_set)}")
        
        # ====== CELL DATA (All cells preserved) ======
        cell_data = {}
        for row_idx in range(1, sheet.max_row + 1):
            row_data = {}
            for col_idx in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell_obj = {}
 
                # Extract styling
                cell_style = extract_complete_cell_style(cell)
                
                # Extract values
                cell_val = extract_complete_cell_value(cell)
                if cell_val:
                    cell_obj.update(cell_val)
 
                # Add style if it exists
                if cell_style:
                    cell_obj["s"] = cell_style
 
                # Store cell (preserve all data)
                if cell_obj:
                    row_data[col_idx - 1] = cell_obj
 
            if row_data:
                cell_data[row_idx - 1] = row_data
 
        # ====== ROW DATA (UniverJS format) ======
        row_data = {}
        for row_idx in range(sheet.max_row):
            row_props = {}
            
            # Get explicit row height
            excel_row_idx = row_idx + 1
            row_dimension = sheet.row_dimensions.get(excel_row_idx)
            
            if row_idx in hidden_rows_set:
                # Hidden rows: set height to 0
                row_props["h"] = 0
                row_props["hd"] = 1  # Hidden flag for UniverJS
            elif row_dimension and hasattr(row_dimension, 'height') and row_dimension.height is not None:
                # Explicit height from Excel
                row_props["h"] = max(row_dimension.height, 1)
            
            if row_props:
                row_data[row_idx] = row_props
 
        # ====== COLUMN DATA (UniverJS format with collapsible hidden columns) ======
        column_data = {}
        for col_idx in range(sheet.max_column):
            col_props = {}
            
            # Get explicit column width
            excel_col_idx = col_idx + 1
            col_letter = get_column_letter(excel_col_idx)
            col_dimension = sheet.column_dimensions.get(col_letter)
            
            if col_idx in hidden_cols_set:
                # Hidden columns: minimal width but still accessible
                col_props["w"] = 2       
                col_props["hd"] = 1  
                col_props["isHidden"] = True  # Additional flag for frontend
            elif col_dimension and hasattr(col_dimension, 'width') and col_dimension.width is not None:
                # Explicit width from Excel (convert Excel units to pixels)
                excel_width = col_dimension.width
                pixel_width = max(excel_width * 7, 20)  # Rough conversion, minimum 20px
                col_props["w"] = pixel_width
            
            if col_props:
                column_data[col_idx] = col_props
 
        # ====== MERGED CELLS ======
        merge_data = []
        for merged in sheet.merged_cells.ranges:
            merge_data.append({
                "startRow": merged.min_row - 1,
                "endRow": merged.max_row - 1,
                "startColumn": merged.min_col - 1,
                "endColumn": merged.max_col - 1
            })
 
        # ====== FREEZE PANES ======
        freeze = get_freeze_info(sheet)
 
        # ====== SHEET JSON (UniverJS compatible) ======
        sheets[sheet_id] = {
            "id": sheet_id,
            "name": sheet.title,
            "tabColor": "",
            "hidden": 0,
            "rowCount": max(sheet.max_row, 100),
            "columnCount": sheet.max_column,
            "zoomRatio": 1,
            "freeze": freeze,
            "scrollTop": 0,
            "scrollLeft": 0,
            "defaultColumnWidth": 73,
            "defaultRowHeight": 23,
            "mergeData": merge_data,
            "cellData": cell_data,
            "rowData": row_data,
            "columnData": column_data,
            "showGridlines": 1,
            "rowHeader": {"width": 46, "hidden": 0},
            "columnHeader": {"height": 20, "hidden": 0},
            "rightToLeft": 0,
            
            # Additional metadata for custom handling
            "_hiddenColumns": sorted(list(hidden_cols_set)),
            "_hiddenRows": sorted(list(hidden_rows_set))
        }
        
        # print(f"Sheet '{sheet.title}' processed: {sheet.max_row} rows, {sheet.max_column} cols")
        # print(f"  - Hidden columns: {len(hidden_cols_set)}")
        # print(f"  - Hidden rows: {len(hidden_rows_set)}")
    
    return {
        "id": str(uuid.uuid4()),
        "name":sheet.title,
        "appVersion": "0.10.2",
        "locale": "enUS",
        "styles": cell_style,
        "sheetOrder": sheet_order,
        "sheets": sheets,
        "resources": [],
    }
